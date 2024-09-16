import frappe
import datetime
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue
from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
from openpyxl.styles.numbers import FORMAT_PERCENTAGE
from frappe.utils import (getdate, cint, add_months, date_diff, add_days,format_date,
    nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime)


@frappe.whitelist()
def download():
    filename = 'Daily Man Power Report'
    test = build_xlsx_response(filename)
    # enqueue(build_xlsx_response, queue='default', timeout=6000, event='build_xlsx_response',filename=filename)
    
def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
         
    ws = wb.create_sheet(sheet_name, 0)
    att_date = getdate(args.att_date)
    department = ''
    
    report_title = ['Manpower Plan vs Actual Report']
    ws.append(report_title)
    week_start_date = getdate(args.att_date) - timedelta(days=getdate(args.att_date).weekday())
    week_end_date = week_start_date + timedelta(days=6)
    att_day = att_date.day
    att_month = att_date.strftime("%B")
    att_year = att_date.year
    
    header1 = ['Department','Manpower Plan  ' + format_date(week_start_date) + '  to  ' + format_date(week_end_date),'','','','','Manpower Status Actual  '+str(att_day)+ str(att_month) + str(att_year),'','','']
    ws.append(header1)
    header2 = ['Shift','Staff','Trainee','Contract','NAPS','Total','Staff','Trainee','Contract','NAPS','Total' ]
    ws.append(header2)
    
    data = get_data(args)

    for row in data:
        ws.append(row)
        
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=11)
    ws.merge_cells(start_row=2,start_column=2,end_row=2,end_column=6) 
    ws.merge_cells(start_row=2,start_column=7,end_row=2,end_column=11) 


    border = Border(left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000'))
    
    if args.dept == 'All Departments':
        for rows in ws.iter_rows(min_row=1, max_row=20, min_col=1, max_col=11):
            for cell in rows:
                cell.border = border
    else:
        for rows in ws.iter_rows(min_row=1, max_row=5, min_col=1, max_col=11):
            for cell in rows:
                cell.border = border
                    
    
    align_center = Alignment(horizontal='center',vertical='center')

    for cell in ws["1:1"]:
        cell.font = Font(bold=True,size=15)
        cell.alignment = align_center 
    
    for cell in ws["2:2"]:
        cell.font = Font(bold=True,size=12)
        cell.alignment = align_center 
        
    
    ws['B2'].alignment = align_center 
    ws['G2'].alignment = align_center   
    
    ws['A1'].fill = PatternFill(fgColor="FFFF00", fill_type = "solid")
    ws['A2'].fill = PatternFill(fgColor="FFA500", fill_type = "solid")
    ws['B2'].fill = PatternFill(fgColor="98FB98", fill_type = "solid")
    ws['G2'].fill = PatternFill(fgColor="00BFFF", fill_type = "solid")
    
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file    

def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'
    

def get_data(args):
    data = []
    staff_count = 0
    tt_count = 0
    cl_count = 0
    naps_count = 0
    att_staff_count = 0
    att_tt_count = 0
    att_cl_count = 0
    att_naps_count = 0
    att_staff_count = 0
    att_tt_count = 0
    att_cl_count = 0
    att_naps_count = 0
    total_over_all_count = 0
    total_over_all_att_count = 0
    total_att_staff_count = 0
    total_att_tt_count = 0
    total_att_tt_count = 0
    total_att_cl_count = 0
    total_att_naps_count = 0
    week_start_date = getdate(args.att_date) - timedelta(days=getdate(args.att_date).weekday())
    week_end_date = week_start_date + timedelta(days=6)
    if args.dept == 'All Departments': 
        row2 = []
        department = frappe.db.get_all('Department', {'docstatus': 'Enabled'}, ['name'], order_by='name asc')
        total_staff_count = 0
        total_tt_count = 0
        total_cl_count = 0
        total_naps_count = 0
        for dept in department:
            if dept.name != 'All Departments':
                row1 = []
                #staff_shift_assignment
                shift_assignment_staff = frappe.db.sql(""" SELECT  e.employee_category, COUNT(*) AS count FROM `tabShift Assignment` sa LEFT JOIN `tabEmployee` e ON sa.employee = e.name AND e.status = 'Active' WHERE sa.start_date BETWEEN %s AND %s AND sa.department = %s AND sa.shift_type = %s AND sa.docstatus = 1  AND e.employee_category = 'Staff'   -- filter by employee category GROUP BY e.employee_category """, (week_start_date, week_end_date, dept.name, args.shift), as_dict=True)
                if shift_assignment_staff:
                    result = shift_assignment_staff[0]
                    staff_count = result['count']
                    total_staff_count += staff_count
                else:
                    staff_count = 0

                #trainee_shift_assignment    
                shift_assignment_tt = frappe.db.sql(""" SELECT  e.employee_category, COUNT(*) AS count FROM `tabShift Assignment` sa LEFT JOIN `tabEmployee` e ON sa.employee = e.name AND e.status = 'Active' WHERE sa.start_date BETWEEN %s AND %s AND sa.department = %s AND sa.shift_type = %s AND sa.docstatus = 1  AND e.employee_category = 'TT'   -- filter by employee category GROUP BY e.employee_category """, (week_start_date, week_end_date, dept.name, args.shift), as_dict=True)
                if shift_assignment_tt:
                    result = shift_assignment_tt[0]
                    tt_count = result['count']
                    total_tt_count += tt_count
                else:
                    tt_count = 0  

                #cl_shift_assignments     
                shift_assignment_cl = frappe.db.sql(""" SELECT  e.employee_category, COUNT(*) AS count FROM `tabShift Assignment` sa LEFT JOIN `tabEmployee` e ON sa.employee = e.name AND e.status = 'Active' WHERE sa.start_date BETWEEN %s AND %s AND sa.department = %s AND sa.shift_type = %s AND sa.docstatus = 1  AND e.employee_category = 'CL'   -- filter by employee category GROUP BY e.employee_category """, (week_start_date, week_end_date, dept.name, args.shift), as_dict=True)
                if shift_assignment_cl:
                    result = shift_assignment_cl[0]
                    cl_count = result['count']
                    total_cl_count += cl_count
                else:
                    cl_count = 0 

                #naps_shift_assignment    
                shift_assignment_naps = frappe.db.sql(""" SELECT  e.employee_category, COUNT(*) AS count FROM `tabShift Assignment` sa LEFT JOIN `tabEmployee` e ON sa.employee = e.name AND e.status = 'Active' WHERE sa.start_date BETWEEN %s AND %s AND sa.department = %s AND sa.shift_type = %s AND sa.docstatus = 1  AND e.employee_category = 'Naps'   -- filter by employee category GROUP BY e.employee_category """, (week_start_date, week_end_date, dept.name, args.shift), as_dict=True)
                if shift_assignment_naps:
                    result = shift_assignment_naps[0]
                    naps_count = result['count']
                    total_naps_count += naps_count
                else:
                    naps_count = 0              
                #staff_attendance
                attendance_count_staff = frappe.db.sql(""" SELECT category, COUNT(*) AS count FROM `tabAttendance`  WHERE attendance_date = %s AND department = %s AND shift = %s and category = 'Staff' GROUP BY category""", (args.att_date,dept.name,args.shift), as_dict=True)
                if attendance_count_staff:
                    result = attendance_count_staff[0]
                    att_staff_count = result['count']
                    total_att_staff_count += att_staff_count
                else:
                    att_staff_count = 0
                #tt_attendance
                attendance_count_staff = frappe.db.sql(""" SELECT category, COUNT(*) AS count FROM `tabAttendance`  WHERE attendance_date = %s AND department = %s AND shift = %s and category = 'TT' GROUP BY category""", (args.att_date,dept.name,args.shift), as_dict=True)
                if attendance_count_staff:
                    result = attendance_count_staff[0]
                    att_tt_count = result['count']
                    total_att_tt_count += att_tt_count
                else:
                    att_tt_count = 0 
                #cl_attendance
                attendance_count_staff = frappe.db.sql(""" SELECT category, COUNT(*) AS count FROM `tabAttendance`  WHERE attendance_date = %s AND department = %s AND shift = %s and category = 'CL' GROUP BY category""", (args.att_date,dept.name,args.shift), as_dict=True)
                if attendance_count_staff:
                    result = attendance_count_staff[0]
                    att_cl_count = result['count']
                    total_att_cl_count += att_cl_count
                else:
                    att_cl_count = 0 
                #naps_attendance
                attendance_count_staff = frappe.db.sql(""" SELECT category, COUNT(*) AS count FROM `tabAttendance`  WHERE attendance_date = %s AND department = %s AND shift = %s and category = 'Naps' GROUP BY category""", (args.att_date,dept.name,args.shift), as_dict=True)
                if attendance_count_staff:
                    result = attendance_count_staff[0]
                    att_naps_count = result['count']
                    total_att_naps_count += att_naps_count
                else:
                    att_naps_count = 0            

                row1 += [dept.name,staff_count,tt_count,cl_count,naps_count,sum([staff_count+tt_count+cl_count]),att_staff_count,att_tt_count,att_cl_count,att_naps_count,sum([att_staff_count+att_tt_count+att_cl_count+att_naps_count])]
                data.append(row1)
        total_over_all_count = sum([total_staff_count+total_tt_count+total_cl_count+total_naps_count]) 
        total_over_all_att_count = sum([total_att_staff_count+total_att_tt_count+total_cl_count+total_att_cl_count])        
        row2 = ['Total',total_staff_count,total_tt_count,total_cl_count,total_naps_count,total_over_all_count,total_att_staff_count,total_att_tt_count,total_att_cl_count,total_att_naps_count,total_over_all_att_count]
        data.append(row2)
                
    else:
        row1 = []
        row3 = []
        staff_count = 0
        tt_count = 0
        cl_count = 0
        naps_count = 0
        att_staff_count = 0
        att_tt_count = 0
        att_cl_count = 0
        att_naps_count = 0

        shift_assignment_staff = frappe.db.sql(""" SELECT  e.employee_category, COUNT(*) AS count FROM `tabShift Assignment` sa LEFT JOIN `tabEmployee` e ON sa.employee = e.name AND e.status = 'Active' WHERE sa.start_date BETWEEN %s AND %s AND sa.department = %s AND sa.shift_type = %s AND sa.docstatus = 1  AND e.employee_category = 'Staff'   -- filter by employee category GROUP BY e.employee_category """, (week_start_date, week_end_date,args.dept,args.shift), as_dict=True)
        if shift_assignment_staff:
            result = shift_assignment_staff[0]
            staff_count = result['count']
        else:
            staff_count = 0

        shift_assignment_staff = frappe.db.sql(""" SELECT  e.employee_category, COUNT(*) AS count FROM `tabShift Assignment` sa LEFT JOIN `tabEmployee` e ON sa.employee = e.name AND e.status = 'Active' WHERE sa.start_date BETWEEN %s AND %s AND sa.department = %s AND sa.shift_type = %s AND sa.docstatus = 1  AND e.employee_category = 'TT'   -- filter by employee category GROUP BY e.employee_category """, (week_start_date, week_end_date,args.dept,args.shift), as_dict=True)
        if shift_assignment_staff:
            result = shift_assignment_staff[0]
            tt_count = result['count']
        else:
            tt_count = 0

        shift_assignment_staff = frappe.db.sql(""" SELECT  e.employee_category, COUNT(*) AS count FROM `tabShift Assignment` sa LEFT JOIN `tabEmployee` e ON sa.employee = e.name AND e.status = 'Active' WHERE sa.start_date BETWEEN %s AND %s AND sa.department = %s AND sa.shift_type = %s AND sa.docstatus = 1  AND e.employee_category = 'CL'   -- filter by employee category GROUP BY e.employee_category """, (week_start_date, week_end_date,args.dept,args.shift), as_dict=True)
        if shift_assignment_staff:
            result = shift_assignment_staff[0]
            cl_count = result['count']
        else:
            cl_count = 0 

        shift_assignment_staff = frappe.db.sql(""" SELECT  e.employee_category, COUNT(*) AS count FROM `tabShift Assignment` sa LEFT JOIN `tabEmployee` e ON sa.employee = e.name AND e.status = 'Active' WHERE sa.start_date BETWEEN %s AND %s AND sa.department = %s AND sa.shift_type = %s AND sa.docstatus = 1  AND e.employee_category = 'Naps'   -- filter by employee category GROUP BY e.employee_category """, (week_start_date, week_end_date,args.dept,args.shift), as_dict=True)
        if shift_assignment_staff:
            result = shift_assignment_staff[0]
            naps_count = result['count']
        else:
            naps_count = 0  

        attendance_count_staff = frappe.db.sql(""" SELECT category, COUNT(*) AS count FROM `tabAttendance`  WHERE attendance_date = %s AND department = %s AND shift = %s and category = 'Staff' GROUP BY category""", (args.att_date,args.dept,args.shift), as_dict=True)
        if attendance_count_staff:
            result = attendance_count_staff[0]
            att_staff_count = result['count'] 

        attendance_count_staff = frappe.db.sql(""" SELECT category, COUNT(*) AS count FROM `tabAttendance`  WHERE attendance_date = %s AND department = %s AND shift = %s and category = 'TT' GROUP BY category""", (args.att_date,args.dept,args.shift), as_dict=True)
        if attendance_count_staff:
            result = attendance_count_staff[0]
            att_tt_count = result['count'] 

        attendance_count_staff = frappe.db.sql(""" SELECT category, COUNT(*) AS count FROM `tabAttendance`  WHERE attendance_date = %s AND department = %s AND shift = %s and category = 'CL' GROUP BY category""", (args.att_date,args.dept,args.shift), as_dict=True)
        if attendance_count_staff:
            result = attendance_count_staff[0]
            att_cl_count = result['count']

        attendance_count_staff = frappe.db.sql(""" SELECT category, COUNT(*) AS count FROM `tabAttendance`  WHERE attendance_date = %s AND department = %s AND shift = %s and category = 'Naps' GROUP BY category""", (args.att_date,args.dept,args.shift), as_dict=True)
        if attendance_count_staff:
            result = attendance_count_staff[0]
            att_naps_count = result['count']                         
             
        row1 = [args.dept,staff_count,tt_count,cl_count,naps_count,sum([staff_count+tt_count+cl_count]),att_staff_count,att_tt_count,att_cl_count,att_naps_count,sum([att_staff_count+att_tt_count+att_cl_count+att_naps_count])]
        row3 = ['Total',staff_count,tt_count,cl_count,naps_count,sum([staff_count+tt_count+cl_count]),att_staff_count,att_tt_count,att_cl_count,att_naps_count,sum([att_staff_count+att_tt_count+att_cl_count+att_naps_count])]
        data.append(row1)
        data.append(row3)                  
    return data            