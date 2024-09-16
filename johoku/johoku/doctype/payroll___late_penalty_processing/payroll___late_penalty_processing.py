# Copyright (c) 2023, TEAMPRO and contributors
# For license information, please see license.txt

import math
import frappe
from frappe.utils.csvutils import read_csv_content
from frappe.utils import (getdate, cint, add_months, date_diff, add_days,
	nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime,datetime,get_first_day,get_last_day,today)
from datetime import datetime
from datetime import timedelta
import locale
from datetime import date
from frappe.model.document import Document
import datetime 
import frappe,erpnext
from frappe.utils import cint
import json
from frappe.utils import date_diff, add_months,today,add_days,add_years,nowdate,flt
from frappe.model.mapper import get_mapped_doc
from frappe.utils.file_manager import get_file
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
import datetime
from datetime import date,datetime,timedelta
import openpyxl
from openpyxl import Workbook
import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
import pandas as pd
from frappe.utils import formatdate
from frappe.utils import now
from erpnext.setup.utils import get_exchange_rate
from frappe.utils import date_diff


class PayrollLatePenaltyProcessing(Document):
	pass

@frappe.whitelist()
def attendance_calc(from_date,to_date):
	employees = frappe.get_all("Employee",{"status":"Active"},["*"])
	for emp in employees:
		attendance = frappe.db.sql("""select name,in_time,shift from `tabAttendance` where employee = '%s' and attendance_date between '%s' and '%s'"""%(emp.name,from_date,to_date),as_dict=True)
		actual_late = 0.0 	
		leave = 0.0
		late = 0.0
		for att in attendance:
			in_time = att['in_time']
			if in_time:
				time_only = timedelta(hours=in_time.hour, minutes=in_time.minute, seconds=in_time.second)
				shift_time = frappe.db.get_value("Shift Type",{'name':att['shift']},['start_time'])
				mins = shift_time + timedelta(minutes=5)
				if  shift_time < time_only < mins:
					actual_late +=1
				elif time_only > mins :
					leave +=1
				if actual_late >= 0:
					at = actual_late
				else:
					at = 0.0
				if at > 3:
					att = at-3
					act_late = att+leave
					late = (leave+att) * 0.5
				else:
					act_late = leave
					late = leave * 0.5
				late_list = actual_late + leave
		ad = frappe.db.sql("""select * from `tabSalary Structure Assignment` where employee = '%s' and docstatus = 1 ORDER BY from_date DESC LIMIT 1  """%(emp.name),as_dict=True)
		if ad:
			days = date_diff(to_date, from_date)
			if late > 0:
				if frappe.db.exists('Late Penalty',{'emp_name':emp.name,'from_date':from_date,'to_date':to_date}):
					adsl = frappe.get_doc('Late Penalty',{'emp_name':emp.name,'from_date':from_date,'to_date':to_date})
					adsl.emp_name = emp.name
					adsl.deduction_days = late
					adsl.actual_late = act_late
					adsl.late_days = late_list 
					adsl.from_date = from_date
					adsl.to_date = to_date
					adsl.late_penalty = (late * ((emp.basic + emp.house_rent_allowance + emp.conveyance_allowance + emp.food_allowance + emp.education_allowance + emp.medical_allowance + emp.mobile_allowance +emp.special_allowance)/(days + 1)))
					adsl.save()
				else:
					adsl = frappe.new_doc("Late Penalty")
					adsl.emp_name = emp.name
					adsl.deduction_days = late
					adsl.actual_late = act_late
					adsl.late_days = late_list 
					adsl.from_date = from_date
					adsl.to_date = to_date
					adsl.late_penalty = (late * ((emp.basic + emp.house_rent_allowance + emp.conveyance_allowance + emp.food_allowance + emp.education_allowance + emp.medical_allowance + emp.mobile_allowance +emp.special_allowance)/(days + 1)))
					adsl.save()
	return "ok"

@frappe.whitelist()
def additional_salary(from_date,to_date):
	employees = frappe.get_all("Employee",{"status":"Active"},["*"])
	for emp in employees:
		lpp = frappe.db.sql("""select * from `tabLate Penalty` where emp_name = '%s' and from_date = '%s' and to_date = '%s'"""%(emp.name,from_date,to_date),as_dict=True)
		if lpp:
			lpp = lpp[0]
			date_str1 = lpp.from_date
			date_obj1 = datetime.strptime(str(date_str1), '%Y-%m-%d').date()
			date_str2 = emp.date_of_joining
			date_obj2 = datetime.strptime(str(date_str2), '%Y-%m-%d').date()
			if date_obj1 > date_obj2:
				payroll_date = date_obj1
			else:
				payroll_date = date_obj2
			if frappe.db.exists('Additional Salary',{'employee':emp.name,'payroll_date':payroll_date}):
				ad = frappe.get_doc('Additional Salary',{'employee':emp.name,'payroll_date':payroll_date})
				ad.employee = emp.name
				ad.salary_component = "Late Penalty"
				ad.type = "Deduction"
				ad.company = emp.company
				ad.amount = lpp.late_penalty
				ad.payroll_date = payroll_date
				ad.save(ignore_permissions=True)
				ad.submit()
				frappe.errprint("Late Penalty Created via Additional Salary")
			else:
				ad = frappe.new_doc('Additional Salary')
				ad.employee = emp.name
				ad.salary_component = "Late Penalty"
				ad.type = "Deduction"
				ad.company = emp.company
				ad.amount = lpp.late_penalty
				ad.payroll_date = payroll_date
				ad.save(ignore_permissions=True)
				ad.submit()
				frappe.errprint("Late Penalty Created via Additional Salary")			
	return "ok"

